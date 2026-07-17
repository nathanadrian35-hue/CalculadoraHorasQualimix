"""
leitor_ponto.py
----------------
Leitura da planilha de ponto eletrônico (Cap. 3).

Responsabilidade única: ler o arquivo XLS/XLSX e produzir dados
estruturados — (mês, ano, list[FuncionarioPlanilha]). Nunca acessa
Config, nunca compara com o cadastro, nunca sugere turno/setor: isso é
responsabilidade de cadastro.py, que consome esta saída.

Estrutura confirmada contra uma exportação real da Qualimix
(dados/RegistroPresença.xls — Sprint 3.5):

    - Uma célula, em qualquer posição da planilha, com texto no formato
      "Data de presença:DD.MM.YYYY~DD.MM.YYYY" — é daí que vem a
      competência (Cap. 3.6), nunca do nome do arquivo. O intervalo
      pode ser parcial (ex.: 13 dias) e, em tese, cruzar dois meses.
    - A partir daí, blocos de exatamente 3 linhas por funcionário
      (Cap. 3.2), sem linha separadora entre blocos:
        Linha 1: rótulo "IDUsuário:" + valor, rótulo "Nome:" + valor,
                 rótulo "Dep.:" + valor (departamento)
        Linha 2: um dia-do-mês (valor numérico real da célula, não a
                 posição da coluna) por coluna
        Linha 3: as batidas daquele dia (texto separado por "\\n"),
                 alinhadas pela mesma coluna da linha de dias
    - Os rótulos de cabeçalho podem ter variação de acentuação (ex.: a
      amostra real grava "IDUsuário:" com "á" em vez de "í" — um
      artefato do sistema de origem) — por isso a comparação usa
      `modelos.normalizar_texto`, não uma string exata.
    - Nenhuma célula de dado (dias/batidas/cabeçalho) está mesclada na
      amostra real — só o título da planilha. Tratado defensivamente
      mesmo assim para .xlsx (openpyxl expõe os intervalos mesclados).
    - `IDUsuário` não é sequencial nem contíguo — é apenas informativo/
      log; a localização do funcionário cadastrado é sempre por nome
      (Cap. 5.7/5.8, `modelos.localizar_funcionario_por_nome`).
"""

from __future__ import annotations

import re
from datetime import date, time
from pathlib import Path
from typing import Any

from logger import get_logger
from modelos import Batida, DiaTrabalho, FuncionarioPlanilha, normalizar_texto

log = get_logger()


class PlanilhaInvalidaError(Exception):
    """Erro bloqueante: a planilha não pôde ser lida ou não tem competência localizável."""


_PADRAO_COMPETENCIA = re.compile(
    r"Data de presen[çc]a\s*:\s*(\d{2})\.(\d{2})\.(\d{4})\s*~\s*(\d{2})\.(\d{2})\.(\d{4})",
    re.IGNORECASE,
)

_ROTULO_ID = "idusuario:"
_ROTULO_NOME = "nome:"
_ROTULO_DEP = "dep.:"


# ---------------------------------------------------------------------------
# Ponto de entrada público
# ---------------------------------------------------------------------------

def ler_planilha(caminho: Path) -> tuple[int, int, list[FuncionarioPlanilha]]:
    """
    Lê a planilha de ponto (.xls ou .xlsx) e retorna (mês, ano da data
    inicial da competência, lista de FuncionarioPlanilha). Roteia por
    extensão.

    O (mês, ano) retornados são um resumo da competência para exibição/
    log — cada dia dentro de FuncionarioPlanilha.dias já carrega sua
    própria data completa e correta, mesmo que a competência cruze o
    fim do mês (ver `_montar_dias`).

    Levanta PlanilhaInvalidaError se o arquivo não puder ser aberto, se
    a competência não for localizada, ou se nenhum funcionário for
    reconhecido.
    """
    extensao = caminho.suffix.lower()
    if extensao == ".xls":
        linhas = _ler_linhas_xls(caminho)
    elif extensao == ".xlsx":
        linhas = _ler_linhas_xlsx(caminho)
    else:
        raise PlanilhaInvalidaError(f"Formato não suportado: {extensao or '(sem extensão)'}")

    data_inicio, data_fim = _extrair_competencia(linhas)
    funcionarios = _extrair_blocos(linhas, data_inicio, data_fim)

    return data_inicio.month, data_inicio.year, funcionarios


# ---------------------------------------------------------------------------
# Leitura bruta por formato
# ---------------------------------------------------------------------------

def _ler_linhas_xls(caminho: Path) -> list[list[Any]]:
    """Lê todas as células da primeira aba de um arquivo .xls, como lista de linhas."""
    import xlrd

    try:
        livro = xlrd.open_workbook(str(caminho))
    except Exception as erro:  # pragma: no cover - proteção contra arquivo corrompido
        raise PlanilhaInvalidaError(f"Não foi possível abrir o arquivo .xls: {erro}") from erro

    if livro.nsheets > 1:
        log.info(
            "Planilha .xls com %d abas — processando apenas a primeira (\"%s\").",
            livro.nsheets, livro.sheet_names()[0],
        )

    aba = livro.sheet_by_index(0)
    return [[aba.cell(r, c).value for c in range(aba.ncols)] for r in range(aba.nrows)]


def _ler_linhas_xlsx(caminho: Path) -> list[list[Any]]:
    """
    Lê todas as células da primeira aba de um arquivo .xlsx, resolvendo
    células mescladas (preenche as posições "vazias" do intervalo com o
    valor da célula-âncora).
    """
    import openpyxl

    try:
        livro = openpyxl.load_workbook(str(caminho), data_only=True)
    except Exception as erro:  # pragma: no cover - proteção contra arquivo corrompido
        raise PlanilhaInvalidaError(f"Não foi possível abrir o arquivo .xlsx: {erro}") from erro

    if len(livro.sheetnames) > 1:
        log.info(
            "Planilha .xlsx com %d abas — processando apenas a primeira (\"%s\").",
            len(livro.sheetnames), livro.sheetnames[0],
        )

    planilha = livro.worksheets[0]

    linhas_brutas = [list(linha) for linha in planilha.iter_rows(values_only=True)]
    largura_maxima = max((len(linha) for linha in linhas_brutas), default=0)
    linhas: list[list[Any]] = [
        linha + [None] * (largura_maxima - len(linha)) for linha in linhas_brutas
    ]

    for intervalo in planilha.merged_cells.ranges:
        linha_ancora, coluna_ancora = intervalo.min_row - 1, intervalo.min_col - 1
        if linha_ancora >= len(linhas) or coluna_ancora >= largura_maxima:
            continue
        valor_ancora = linhas[linha_ancora][coluna_ancora]
        for r in range(intervalo.min_row - 1, min(intervalo.max_row, len(linhas))):
            for c in range(intervalo.min_col - 1, min(intervalo.max_col, largura_maxima)):
                linhas[r][c] = valor_ancora

    return linhas


# ---------------------------------------------------------------------------
# Competência (Cap. 3.6)
# ---------------------------------------------------------------------------

def _extrair_competencia(linhas: list[list[Any]]) -> tuple[date, date]:
    """
    Localiza, em qualquer célula da planilha, o texto "Data de
    presença:DD.MM.YYYY~DD.MM.YYYY" e retorna (data_inicio, data_fim).
    Não assume posição fixa — varre todas as células até encontrar.
    """
    for linha in linhas:
        for valor in linha:
            if not isinstance(valor, str):
                continue
            correspondencia = _PADRAO_COMPETENCIA.search(valor)
            if correspondencia:
                d1, m1, a1, d2, m2, a2 = (int(g) for g in correspondencia.groups())
                try:
                    return date(a1, m1, d1), date(a2, m2, d2)
                except ValueError as erro:
                    log.error("Datas de competência inválidas na planilha: %s", erro)
                    raise PlanilhaInvalidaError(f"Datas de competência inválidas: {erro}") from erro

    log.error('Competência não localizada — nenhuma célula com "Data de presença:..." encontrada.')
    raise PlanilhaInvalidaError(
        'Não foi possível localizar a competência (célula "Data de presença:...").'
    )


# ---------------------------------------------------------------------------
# Blocos de funcionário (Cap. 3.2/3.3)
# ---------------------------------------------------------------------------

def _indice_do_rotulo(linha: list[Any], rotulo_normalizado: str) -> int | None:
    """Retorna a coluna onde `rotulo_normalizado` aparece na linha, ou None."""
    for indice, valor in enumerate(linha):
        if isinstance(valor, str) and normalizar_texto(valor) == rotulo_normalizado:
            return indice
    return None


def _valor_apos_rotulo(linha: list[Any], indice_rotulo: int) -> Any:
    """Retorna o primeiro valor não vazio após a coluna do rótulo, na mesma linha."""
    for indice in range(indice_rotulo + 1, len(linha)):
        valor = linha[indice]
        if valor not in ("", None):
            return valor
    return None


def _extrair_blocos(
    linhas: list[list[Any]], data_inicio: date, data_fim: date,
) -> list[FuncionarioPlanilha]:
    """
    Varre as linhas procurando blocos de 3 linhas por funcionário
    (Cap. 3.2): cabeçalho (IDUsuário/Nome/Dep.), dias, batidas.
    Tolerante a linhas vazias entre blocos e a blocos incompletos —
    nunca interrompe a leitura dos demais (Cap. 9.1).
    """
    funcionarios: list[FuncionarioPlanilha] = []
    total = len(linhas)
    indice = 0

    while indice < total:
        indice_id = _indice_do_rotulo(linhas[indice], _ROTULO_ID)
        if indice_id is None:
            indice += 1
            continue

        id_planilha = _valor_apos_rotulo(linhas[indice], indice_id)

        indice_nome = _indice_do_rotulo(linhas[indice], _ROTULO_NOME)
        nome_planilha = (
            _valor_apos_rotulo(linhas[indice], indice_nome) if indice_nome is not None else None
        )

        indice_dep = _indice_do_rotulo(linhas[indice], _ROTULO_DEP)
        departamento = (
            _valor_apos_rotulo(linhas[indice], indice_dep) if indice_dep is not None else None
        )

        if not nome_planilha:
            log.warning("Bloco na linha %d sem Nome legível — ignorado.", indice + 1)
            indice += 1
            continue

        linha_dias = linhas[indice + 1] if indice + 1 < total else []
        linha_batidas = linhas[indice + 2] if indice + 2 < total else []

        dias = _montar_dias(linha_dias, linha_batidas, data_inicio, data_fim)

        funcionarios.append(FuncionarioPlanilha(
            id_planilha=str(id_planilha).strip() if id_planilha is not None else "",
            nome_planilha=str(nome_planilha).strip(),
            departamento=str(departamento).strip() if departamento else "",
            dias=dias,
        ))

        indice += 3  # bloco sempre ocupa 3 linhas (Cap. 3.2)

    funcionarios = _mesclar_repetidos(funcionarios)

    if not funcionarios:
        log.error("Nenhum funcionário reconhecido na planilha.")
        raise PlanilhaInvalidaError("Nenhum funcionário reconhecido na planilha.")

    return funcionarios


def _mesclar_repetidos(funcionarios: list[FuncionarioPlanilha]) -> list[FuncionarioPlanilha]:
    """
    Mescla blocos repetidos (mesmo bloco do mesmo funcionário aparecendo
    mais de uma vez na planilha) em um único FuncionarioPlanilha, somando
    os dias. Chaveia pelo `id_planilha` ("IDUsuário") quando presente —
    identificador principal (Cap. 5.8) — e só cai para o nome normalizado
    quando ele vier vazio. Isso é o que garante que dois funcionários
    REAIS com o mesmo nome (homônimos, IDUsuário diferente) cheguem como
    dois registros distintos ao Painel de Revisão, em vez de serem
    fundidos incondicionalmente como antes. Planilhas normalmente não
    repetem funcionários (não há nenhum caso na amostra real), mas o
    parser não deve travar nem duplicar se isso ocorrer.
    """
    por_chave: dict[str, FuncionarioPlanilha] = {}
    ordem: list[str] = []
    for funcionario in funcionarios:
        chave = funcionario.id_planilha.strip() or normalizar_texto(funcionario.nome_planilha)
        if chave in por_chave:
            log.warning(
                "Funcionário repetido na planilha: \"%s\" — batidas mescladas em um único "
                "registro.",
                funcionario.nome_planilha,
            )
            por_chave[chave].dias.extend(funcionario.dias)
        else:
            por_chave[chave] = funcionario
            ordem.append(chave)
    return [por_chave[chave] for chave in ordem]


# ---------------------------------------------------------------------------
# Dias e batidas (Cap. 3.4/3.5/3.7)
# ---------------------------------------------------------------------------

def _montar_dias(
    linha_dias: list[Any], linha_batidas: list[Any], data_inicio: date, data_fim: date,
) -> list[DiaTrabalho]:
    """
    Constrói a lista de DiaTrabalho a partir da linha de dias e da linha
    de batidas do bloco, alinhadas por coluna (Cap. 3.4/3.5).

    Suporta intervalos que cruzam meses: detecta quando o valor do dia
    "regride" em relação ao anterior (ex.: 30 seguido de 1) e avança o
    mês, com rollover de ano em dezembro -> janeiro, usando data_inicio
    como âncora do primeiro mês.
    """
    dias: list[DiaTrabalho] = []
    ano_atual, mes_atual = data_inicio.year, data_inicio.month
    dia_anterior: int | None = None

    for coluna, valor_dia in enumerate(linha_dias):
        if not isinstance(valor_dia, (int, float)):
            continue
        numero_dia = int(valor_dia)
        if not (1 <= numero_dia <= 31):
            log.warning("Valor de dia inválido na coluna %d: %r — ignorado.", coluna, valor_dia)
            continue

        if dia_anterior is not None and numero_dia <= dia_anterior:
            mes_atual += 1
            if mes_atual > 12:
                mes_atual = 1
                ano_atual += 1
        dia_anterior = numero_dia

        try:
            data_completa = date(ano_atual, mes_atual, numero_dia)
        except ValueError:
            log.warning(
                "Data inválida montada a partir do dia %d (%04d-%02d) — dia ignorado.",
                numero_dia, ano_atual, mes_atual,
            )
            continue

        if not (data_inicio <= data_completa <= data_fim):
            log.warning(
                "Data %s fora do intervalo da competência (%s a %s) — mantida mesmo assim.",
                data_completa.strftime("%d/%m/%Y"),
                data_inicio.strftime("%d/%m/%Y"), data_fim.strftime("%d/%m/%Y"),
            )

        texto_batidas = linha_batidas[coluna] if coluna < len(linha_batidas) else ""
        batidas = _parse_batidas(texto_batidas)

        dias.append(DiaTrabalho(
            data=data_completa,
            dia_semana=_nome_dia_semana(data_completa),
            batidas=batidas,
        ))

    return dias


def _nome_dia_semana(data: date) -> str:
    """Nome do dia da semana em português (Cap. 3.7), calculado — nunca lido da planilha."""
    from constantes import nome_dia_semana
    return nome_dia_semana(data.weekday())


def _parse_batidas(texto: Any) -> list[Batida]:
    """
    Separa o texto de uma célula de batidas em uma lista de Batida
    (Cap. 3.5), ignorando trechos vazios e horários ilegíveis (a
    validação estrutural, não a leitura, é quem registra a pendência).
    """
    if not isinstance(texto, str) or not texto.strip():
        return []

    batidas: list[Batida] = []
    for trecho in texto.split("\n"):
        trecho = trecho.strip()
        if not trecho:
            continue
        horario = _parse_horario(trecho)
        if horario is not None:
            batidas.append(Batida(horario=horario))
        else:
            log.warning("Horário ilegível em célula de batida: %r — ignorado.", trecho)

    return batidas


def _parse_horario(texto: str) -> time | None:
    """Converte um texto "HH:MM" em datetime.time. Retorna None se ilegível."""
    partes = texto.split(":")
    if len(partes) != 2:
        return None
    try:
        hora, minuto = int(partes[0]), int(partes[1])
        return time(hour=hora, minute=minuto)
    except ValueError:
        return None
