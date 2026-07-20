"""
relatorio.py
------------
Geração de Relatórios (Cap. 11) — Sprint 5.

Responsabilidade única (Cap. 18): toda a lógica de relatório mora
aqui. Consome exclusivamente o que o Motor de Cálculo (calculadora.py,
Cap. 6) já produziu — `ResultadoProcessamento`, `DiaTrabalho.resultado`
(`ResultadoDia`) e `ResumoMensalFuncionario`. Nenhuma regra de cálculo
é reimplementada: tudo aqui é agregação (soma/contagem) e formatação
de valores que o Motor já calculou — nenhuma hora é derivada de novo a
partir de batidas.

A planilha original nunca é tocada (Cap. 11.1/11.2) — só é lida, lá
atrás, por leitor_ponto.py. Este módulo só escreve arquivos novos.
"""

from __future__ import annotations

import copy
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Protocol

from calculadora import montar_resumo_mensal
from config import Config, setor_de_dict, turno_de_dict
from constantes import (
    APP_NOME,
    DESENVOLVEDOR,
    Situacao,
    StatusFuncionario,
    VERSAO,
    formatar_minutos,
)
from logger import get_logger
from modelos import (
    ArquivoHistorico,
    CompetenciaHistorico,
    DadosRelatorio,
    DiaTrabalho,
    EstatisticasCompetencia,
    Funcionario,
    ResultadoProcessamento,
    ResumoGeralCompetencia,
    ResumoIndividualRelatorio,
    Setor,
    Turno,
)

log = get_logger()

MENSAGEM_BLOQUEIO = (
    "Existem pendências que precisam ser resolvidas antes da emissão do relatório."
)

_RODAPE = (
    f"Este relatório foi gerado automaticamente pelo {APP_NOME}.",
    "A planilha oficial exportada pelo relógio de ponto não foi modificada.",
    f"Desenvolvido por {DESENVOLVEDOR}.",
)


# ---------------------------------------------------------------------------
# Bloqueio por pendências (Cap. 9.1 / 11.11)
# ---------------------------------------------------------------------------

class RelatorioBloqueadoError(Exception):
    """
    Levantada ao tentar gerar um relatório com pendências em aberto
    (Cap. 9.1/11.11). Verificada dentro das próprias funções de
    geração (não só na tela) — nenhum chamador consegue burlar a
    regra chamando `relatorio.py` diretamente.
    """


def existem_pendencias_abertas(resultado: ResultadoProcessamento) -> bool:
    """Cap. 9.1/11.11: True se houver qualquer pendência ainda não resolvida."""
    return any(not p.resolvida for p in resultado.pendencias)


def _garantir_sem_bloqueio(resultado: ResultadoProcessamento) -> None:
    if existem_pendencias_abertas(resultado):
        raise RelatorioBloqueadoError(MENSAGEM_BLOQUEIO)


# ---------------------------------------------------------------------------
# Filtro (Cap. 12.8: Funcionário/Setor/Turno/Cargo/Status)
# ---------------------------------------------------------------------------

def filtrar_resultado(
    resultado: ResultadoProcessamento,
    funcionario_id: str | None = None,
    setor_id: str | None = None,
    turno_id: str | None = None,
    cargo: str | None = None,
    status: StatusFuncionario | None = None,
) -> ResultadoProcessamento:
    """
    Recorta um `ResultadoProcessamento` para um subconjunto de
    funcionários (Cap. 12.8) — nenhum valor é recalculado, só
    filtrado. `funcionario_id` restringe a um único funcionário
    (modo "Individual"); os demais parâmetros combinam entre si
    (modo "Todos" com filtros).
    """
    funcionarios = [
        f for f in resultado.funcionarios_processados
        if (funcionario_id is None or f.id == funcionario_id)
        and (setor_id is None or f.setor_id == setor_id)
        and (turno_id is None or f.turno_id == turno_id)
        and (cargo is None or f.cargo == cargo)
        and (status is None or f.status == status)
    ]
    ids = {f.id for f in funcionarios}
    return ResultadoProcessamento(
        funcionarios_processados=funcionarios,
        pendencias=[p for p in resultado.pendencias if p.id_funcionario in ids],
        resumos_mensais=[r for r in resultado.resumos_mensais if r.funcionario_id in ids],
        estatisticas=dict(resultado.estatisticas),
    )


# ---------------------------------------------------------------------------
# Filtro por período (Cap. novo, v2.0) — um dia, uma semana, quinzena,
# intervalo personalizado ou o mês completo, sem depender do fechamento
# da competência (Etapa 8)
# ---------------------------------------------------------------------------

def filtrar_por_periodo(
    resultado: ResultadoProcessamento, data_inicio: date, data_fim: date,
) -> ResultadoProcessamento:
    """
    Recorta um `ResultadoProcessamento` para um intervalo de datas
    (Cap. novo, v2.0). Nenhum valor é recalculado: cada `Funcionario`
    do resultado filtrado é uma cópia rasa com `.dias` recortado ao
    intervalo — os objetos originais (e a Competência persistida)
    nunca são alterados. `resumos_mensais`/`estatisticas` são
    recompostos sobre o subconjunto, reaproveitando a mesma agregação
    pura do Motor (`calculadora.montar_resumo_mensal`) — nenhuma
    fórmula nova.
    """
    funcionarios_recortados: list[Funcionario] = []
    for funcionario in resultado.funcionarios_processados:
        funcionario_copia = copy.copy(funcionario)
        funcionario_copia.dias = [
            dia for dia in funcionario.dias if data_inicio <= dia.data <= data_fim
        ]
        funcionarios_recortados.append(funcionario_copia)

    ids = {f.id for f in funcionarios_recortados}
    pendencias_no_periodo = [
        p for p in resultado.pendencias
        if p.id_funcionario in ids and (p.data is None or data_inicio <= p.data <= data_fim)
    ]

    return ResultadoProcessamento(
        funcionarios_processados=funcionarios_recortados,
        pendencias=pendencias_no_periodo,
        resumos_mensais=[montar_resumo_mensal(f) for f in funcionarios_recortados],
        estatisticas=dict(resultado.estatisticas),
    )


# ---------------------------------------------------------------------------
# Filtros adicionais (Cap. novo, v2.0 — Etapa 9): Situação, Pendências,
# Horas Extras, Banco de Horas (sinônimo de Saldo, Cap. 10.4)
# ---------------------------------------------------------------------------

def filtrar_por_atributos(
    resultado: ResultadoProcessamento,
    situacao: Situacao | None = None,
    com_pendencia: bool | None = None,
    com_horas_extras: bool | None = None,
    banco_horas: str | None = None,
) -> ResultadoProcessamento:
    """
    Filtros por atributo do funcionário (Cap. novo, v2.0), combináveis
    entre si e com `filtrar_resultado`/`filtrar_por_periodo`: nenhum
    valor é recalculado, só filtrado sobre o que o Motor/`resumos_mensais`
    já produziram.

      - `situacao`: mantém quem tem pelo menos 1 dia com essa Situação
        no período considerado.
      - `com_pendencia`: True = só quem tem pendência em aberto; False
        = só quem não tem.
      - `com_horas_extras`: True = só quem tem horas extras > 0 no
        total do período.
      - `banco_horas`: "positivo" / "negativo" / "zerado" — Banco de
        Horas é o mesmo Saldo já calculado (Cap. 10.4), sem nenhuma
        regra de acúmulo entre competências.
    """
    resumos_por_id = {r.funcionario_id: r for r in resultado.resumos_mensais}
    pendencias_abertas_por_id: dict[str, int] = {}
    for pendencia in resultado.pendencias:
        if not pendencia.resolvida:
            pendencias_abertas_por_id[pendencia.id_funcionario] = (
                pendencias_abertas_por_id.get(pendencia.id_funcionario, 0) + 1
            )

    def combina(funcionario: Funcionario) -> bool:
        if situacao is not None and not any(
            dia.resultado.situacao == situacao for dia in funcionario.dias
        ):
            return False
        if com_pendencia is not None:
            tem_pendencia = pendencias_abertas_por_id.get(funcionario.id, 0) > 0
            if tem_pendencia != com_pendencia:
                return False
        if com_horas_extras is not None:
            resumo = resumos_por_id.get(funcionario.id)
            tem_extras = bool(resumo and resumo.horas_extras_min > 0)
            if tem_extras != com_horas_extras:
                return False
        if banco_horas is not None:
            resumo_saldo = resumos_por_id.get(funcionario.id)
            saldo = resumo_saldo.saldo_final_min if resumo_saldo else 0
            if banco_horas == "positivo" and saldo <= 0:
                return False
            if banco_horas == "negativo" and saldo >= 0:
                return False
            if banco_horas == "zerado" and saldo != 0:
                return False
        return True

    funcionarios = [f for f in resultado.funcionarios_processados if combina(f)]
    ids = {f.id for f in funcionarios}
    return ResultadoProcessamento(
        funcionarios_processados=funcionarios,
        pendencias=[p for p in resultado.pendencias if p.id_funcionario in ids],
        resumos_mensais=[r for r in resultado.resumos_mensais if r.funcionario_id in ids],
        estatisticas=dict(resultado.estatisticas),
    )


# ---------------------------------------------------------------------------
# Agregações (Cap. 11.8/11.9/11.10) — somam/contam o que o Motor já calculou
# ---------------------------------------------------------------------------

def montar_resumo_individual(
    funcionario: Funcionario,
    resultado: ResultadoProcessamento,
    setores: list[Setor],
    turnos: list[Turno],
) -> ResumoIndividualRelatorio:
    """
    Resumo de um funcionário para o Relatório Individual (Cap. 11.8).
    Reaproveita o `ResumoMensalFuncionario` que o Motor já calculou
    (trabalhadas/extras/negativas/saldo); soma independentemente
    apenas o que falta nele — jornada prevista total e o
    desdobramento de pendências resolvidas/existentes — sobre dados
    já calculados por dia (`DiaTrabalho.resultado`/`.pendencia`).
    Nenhuma conta nova é feita.
    """
    resumo_motor = next(
        (r for r in resultado.resumos_mensais if r.funcionario_id == funcionario.id), None,
    )
    horas_previstas_min = sum(dia.resultado.jornada_prevista_min for dia in funcionario.dias)
    resolvidas = sum(
        1 for dia in funcionario.dias if dia.pendencia is not None and dia.pendencia.resolvida
    )
    existentes = sum(
        1 for dia in funcionario.dias
        if dia.pendencia is not None and not dia.pendencia.resolvida
    )
    setor_nome = next((s.nome for s in setores if s.id == funcionario.setor_id), "—")
    turno_nome = next((t.nome for t in turnos if t.id == funcionario.turno_id), "—")

    return ResumoIndividualRelatorio(
        funcionario_id=funcionario.id,
        nome=funcionario.nome_completo,
        cargo=funcionario.cargo or "—",
        setor_nome=setor_nome,
        turno_nome=turno_nome,
        horas_previstas_min=horas_previstas_min,
        horas_trabalhadas_min=resumo_motor.horas_trabalhadas_min if resumo_motor else 0,
        horas_extras_min=resumo_motor.horas_extras_min if resumo_motor else 0,
        horas_negativas_min=resumo_motor.horas_negativas_min if resumo_motor else 0,
        saldo_final_min=resumo_motor.saldo_final_min if resumo_motor else 0,
        quantidade_pendencias_resolvidas=resolvidas,
        quantidade_pendencias_existentes=existentes,
    )


def montar_resumo_geral(
    resultado: ResultadoProcessamento, competencia_texto: str,
) -> ResumoGeralCompetencia:
    """Cap. 11.9 — soma o que o Motor já calculou para cada funcionário (`resumos_mensais`)."""
    horas_previstas_min = sum(
        dia.resultado.jornada_prevista_min
        for funcionario in resultado.funcionarios_processados
        for dia in funcionario.dias
    )
    funcionarios_com_pendencia = len({
        p.id_funcionario for p in resultado.pendencias if not p.resolvida
    })

    return ResumoGeralCompetencia(
        competencia_texto=competencia_texto,
        funcionarios_processados=len(resultado.funcionarios_processados),
        funcionarios_com_pendencia=funcionarios_com_pendencia,
        total_pendencias=len(resultado.pendencias),
        horas_previstas_min=horas_previstas_min,
        horas_trabalhadas_min=sum(r.horas_trabalhadas_min for r in resultado.resumos_mensais),
        horas_extras_min=sum(r.horas_extras_min for r in resultado.resumos_mensais),
        horas_negativas_min=sum(r.horas_negativas_min for r in resultado.resumos_mensais),
        saldo_geral_min=sum(r.saldo_final_min for r in resultado.resumos_mensais),
    )


def montar_estatisticas(resultado: ResultadoProcessamento) -> EstatisticasCompetencia:
    """
    Cap. 11.10 — contagens sobre o que o Motor já produziu
    (`pendencias`/`funcionarios_processados`).
    """
    pendencias_por_tipo = dict(Counter(p.tipo for p in resultado.pendencias))
    distribuicao_justificativas = dict(Counter(
        p.justificativa for p in resultado.pendencias if p.justificativa
    ))
    funcionarios_ativos = sum(
        1 for f in resultado.funcionarios_processados if f.status == StatusFuncionario.ATIVO
    )
    funcionarios_sem_batidas = sum(
        1 for f in resultado.funcionarios_processados
        if not any(dia.batidas for dia in f.dias)
    )
    total_dias_processados = sum(len(f.dias) for f in resultado.funcionarios_processados)

    return EstatisticasCompetencia(
        funcionarios_ativos=funcionarios_ativos,
        funcionarios_processados=len(resultado.funcionarios_processados),
        funcionarios_sem_batidas=funcionarios_sem_batidas,
        pendencias_por_tipo=pendencias_por_tipo,
        distribuicao_justificativas=distribuicao_justificativas,
        total_dias_processados=total_dias_processados,
    )


def montar_dados_relatorio(
    resultado: ResultadoProcessamento,
    config: Config,
    arquivo_original: str,
    competencia_texto: str,
) -> DadosRelatorio:
    """
    Ponto de entrada público: monta o pacote completo de dados que
    qualquer exportador (Cap. 11.12) precisa, a partir do
    `ResultadoProcessamento` do Motor e da Config já carregada.
    """
    setores = [setor_de_dict(dados) for dados in config.setores.get("setores", [])]
    turnos = [turno_de_dict(dados) for dados in config.configuracoes.get("turnos", [])]
    agora = datetime.now()

    return DadosRelatorio(
        resultado=resultado,
        resumo_geral=montar_resumo_geral(resultado, competencia_texto),
        estatisticas=montar_estatisticas(resultado),
        nome_empresa=config.nome_empresa,
        logo_caminho=config.empresa.get("logo_caminho", ""),
        competencia_texto=competencia_texto,
        arquivo_original=arquivo_original,
        data_processamento=agora.strftime("%d/%m/%Y"),
        hora_processamento=agora.strftime("%H:%M:%S"),
        setores=setores,
        turnos=turnos,
    )


# ---------------------------------------------------------------------------
# Auxiliares de formatação de um dia (Cap. 11.4 Aba 1 / 11.8)
# ---------------------------------------------------------------------------

def _texto_batidas(dia: DiaTrabalho) -> str:
    """Todas as batidas do dia, na ordem em que ocorreram (Cap. 11.4 Aba 1)."""
    return ", ".join(batida.horario.strftime("%H:%M") for batida in dia.batidas)


def _justificativa_do_dia(dia: DiaTrabalho) -> str:
    return dia.pendencia.justificativa if dia.pendencia is not None else ""


def _observacao_do_dia(dia: DiaTrabalho) -> str:
    if dia.pendencia is not None and dia.pendencia.observacoes:
        return dia.pendencia.observacoes
    return dia.observacoes


def _nome_turno(funcionario: Funcionario, turnos: list[Turno]) -> str:
    return next((t.nome for t in turnos if t.id == funcionario.turno_id), "—")


def _nome_setor(funcionario: Funcionario, setores: list[Setor]) -> str:
    return next((s.nome for s in setores if s.id == funcionario.setor_id), "—")


# ---------------------------------------------------------------------------
# Cabeçalho / rodapé (Cap. 11.5/11.6)
# ---------------------------------------------------------------------------

def _aplicar_cabecalho(aba, dados: DadosRelatorio, titulo_aba: str) -> None:
    """Escreve as linhas de cabeçalho (Cap. 11.5) no topo da aba."""
    aba.append((APP_NOME,))
    aba.append((titulo_aba,))
    aba.append((f"Empresa: {dados.nome_empresa or '—'}",))
    aba.append((f"Competência: {dados.competencia_texto}",))
    aba.append((f"Processado em: {dados.data_processamento} {dados.hora_processamento}",))
    aba.append(())

    if dados.logo_caminho:
        _inserir_logo(aba, dados.logo_caminho)


def _inserir_logo(aba, logo_caminho: str) -> None:
    """Insere a logo no canto da aba (Cap. 11.5); falha silenciosa se não for possível."""
    try:
        from openpyxl.drawing.image import Image as ImagemExcel

        caminho = Path(logo_caminho)
        if not caminho.is_absolute():
            from config import BASE_DIR
            caminho = BASE_DIR / logo_caminho
        if not caminho.exists():
            return
        imagem = ImagemExcel(str(caminho))
        imagem.width, imagem.height = 60, 60
        aba.add_image(imagem, "E1")
    except Exception as erro:  # pragma: no cover - proteção contra imagem inválida
        log.warning("Não foi possível inserir a logo no relatório: %s", erro)


def _aplicar_rodape(aba) -> None:
    """Cap. 11.6 — texto fixo ao final da aba e no rodapé de impressão."""
    aba.append(())
    for linha in _RODAPE:
        aba.append((linha,))
    aba.oddFooter.center.text = " | ".join(_RODAPE)


# ---------------------------------------------------------------------------
# Formatação profissional para conferência/impressão (Cap. 11, item 6 da
# Sprint 5) — formata apenas células já escritas pelos `aba.append(...)`
# acima; nenhum valor é alterado, recalculado ou reordenado aqui.
# ---------------------------------------------------------------------------

def _estilizar_titulo(aba, num_colunas: int) -> None:
    """Cap. 11.5 — título do relatório em destaque, mesclado sobre a largura da tabela."""
    from openpyxl.styles import Font

    aba.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E78")
    aba.cell(row=2, column=1).font = Font(bold=True, size=12)
    for linha in (3, 4, 5):
        aba.cell(row=linha, column=1).font = Font(italic=True, color="595959")
    if num_colunas > 1:
        for linha in (1, 2):
            aba.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=num_colunas)


def _estilizar_tabela(aba, linha_cabecalho: int, linha_final: int, num_colunas: int) -> None:
    """
    Formatação profissional da tabela de dados (Cap. 11): cabeçalho em
    destaque com painel congelado (fica visível ao rolar a planilha),
    bordas finas e listras suaves para facilitar a conferência, e
    largura de coluna ajustada ao conteúdo. Só formata a aparência das
    células já escritas — nenhum valor é alterado.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if linha_final < linha_cabecalho:
        return

    preenchimento_cabecalho = PatternFill(fill_type="solid", fgColor="1F4E78")
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    lado = Side(style="thin", color="BFBFBF")
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)
    listra = PatternFill(fill_type="solid", fgColor="F2F2F2")

    for coluna in range(1, num_colunas + 1):
        celula = aba.cell(row=linha_cabecalho, column=coluna)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
    aba.row_dimensions[linha_cabecalho].height = 26

    for linha in range(linha_cabecalho + 1, linha_final + 1):
        listrada = (linha - linha_cabecalho) % 2 == 0
        for coluna in range(1, num_colunas + 1):
            celula = aba.cell(row=linha, column=coluna)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")
            if listrada:
                celula.fill = listra

    aba.freeze_panes = aba.cell(row=linha_cabecalho + 1, column=1).coordinate

    larguras: dict[int, int] = {}
    for linha in aba.iter_rows(min_row=linha_cabecalho, max_row=linha_final):
        for celula in linha:
            if celula.value is None:
                continue
            larguras[celula.column] = max(larguras.get(celula.column, 0), len(str(celula.value)))
    for coluna, largura in larguras.items():
        aba.column_dimensions[get_column_letter(coluna)].width = min(largura + 2, 40)


def _estilizar_pares_chave_valor(aba, linha_inicio: int, linha_fim: int) -> None:
    """
    Formatação leve para blocos de "rótulo: valor" (cabeçalho do
    Relatório Individual, Resumo do Funcionário, Informações do
    Processamento): rótulo em negrito e largura de coluna ajustada.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    largura_a = 0
    largura_b = 0
    for linha in range(linha_inicio, linha_fim + 1):
        celula_a = aba.cell(row=linha, column=1)
        celula_b = aba.cell(row=linha, column=2)
        if celula_a.value is not None:
            celula_a.font = Font(bold=True)
            largura_a = max(largura_a, len(str(celula_a.value)))
        if celula_b.value is not None:
            largura_b = max(largura_b, len(str(celula_b.value)))

    aba.column_dimensions[get_column_letter(1)].width = min(largura_a + 2, 40)
    aba.column_dimensions[get_column_letter(2)].width = min(max(largura_b + 2, 12), 60)


def _preparar_impressao(aba, num_colunas: int, linha_cabecalho: int | None) -> None:
    """
    Cap. 11, item 6 — layout de impressão: paisagem, ajustada à
    largura da página, com a linha de cabeçalho da tabela repetida em
    cada página impressa (quando existir uma tabela na aba).
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    aba.page_setup.orientation = "landscape"
    aba.page_setup.fitToWidth = 1
    aba.page_setup.fitToHeight = 0
    aba.sheet_properties.pageSetUpPr.fitToPage = True
    aba.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)
    if linha_cabecalho is not None:
        aba.print_title_rows = f"{linha_cabecalho}:{linha_cabecalho}"
    ultima_coluna = get_column_letter(max(num_colunas, 1))
    aba.print_area = f"A1:{ultima_coluna}{aba.max_row}"


# ---------------------------------------------------------------------------
# Relatório Geral — 4 abas (Cap. 11.4)
# ---------------------------------------------------------------------------

_CABECALHO_DIARIO = (
    "Funcionário", "Data", "Dia da Semana", "Turno", "Batidas", "Jornada Prevista",
    "Horas Trabalhadas", "Horas Extras", "Horas Negativas", "Saldo", "Situação",
    "Justificativa", "Observações",
)

_CABECALHO_RESUMO_MENSAL = (
    "Funcionário", "Dias Trabalhados", "Horas Trabalhadas", "Horas Extras",
    "Horas Negativas", "Saldo Final", "Quantidade de Pendências",
)

_CABECALHO_PENDENCIAS = (
    "Funcionário", "Data", "Tipo da Pendência", "Descrição", "Status", "Observações",
)


def _montar_aba_relatorio_diario(aba, dados: DadosRelatorio) -> None:
    aba.title = "Relatório Diário"
    _aplicar_cabecalho(aba, dados, "Relatório Diário")
    aba.append(_CABECALHO_DIARIO)
    linha_cabecalho = aba.max_row

    for funcionario in sorted(
        dados.resultado.funcionarios_processados, key=lambda f: f.nome_completo,
    ):
        turno_nome = _nome_turno(funcionario, dados.turnos)
        for dia in sorted(funcionario.dias, key=lambda d: d.data):
            aba.append((
                funcionario.nome_completo,
                dia.data.strftime("%d/%m/%Y"),
                dia.dia_semana,
                turno_nome,
                _texto_batidas(dia),
                formatar_minutos(dia.resultado.jornada_prevista_min),
                formatar_minutos(dia.resultado.horas_trabalhadas_min),
                formatar_minutos(dia.resultado.horas_extras_min),
                formatar_minutos(dia.resultado.horas_negativas_min),
                formatar_minutos(dia.resultado.saldo_min),
                dia.resultado.situacao.value,
                _justificativa_do_dia(dia),
                _observacao_do_dia(dia),
            ))

    linha_final = aba.max_row
    _aplicar_rodape(aba)
    _estilizar_titulo(aba, len(_CABECALHO_DIARIO))
    _estilizar_tabela(aba, linha_cabecalho, linha_final, len(_CABECALHO_DIARIO))
    _preparar_impressao(aba, len(_CABECALHO_DIARIO), linha_cabecalho)


def _montar_aba_resumo_mensal(aba, dados: DadosRelatorio) -> None:
    aba.title = "Resumo Mensal"
    _aplicar_cabecalho(aba, dados, "Resumo Mensal")
    aba.append(_CABECALHO_RESUMO_MENSAL)
    linha_cabecalho = aba.max_row

    for resumo in sorted(dados.resultado.resumos_mensais, key=lambda r: r.nome):
        aba.append((
            resumo.nome,
            resumo.dias_trabalhados,
            formatar_minutos(resumo.horas_trabalhadas_min),
            formatar_minutos(resumo.horas_extras_min),
            formatar_minutos(resumo.horas_negativas_min),
            formatar_minutos(resumo.saldo_final_min),
            resumo.quantidade_pendencias,
        ))

    linha_final = aba.max_row
    _aplicar_rodape(aba)
    _estilizar_titulo(aba, len(_CABECALHO_RESUMO_MENSAL))
    _estilizar_tabela(aba, linha_cabecalho, linha_final, len(_CABECALHO_RESUMO_MENSAL))
    _preparar_impressao(aba, len(_CABECALHO_RESUMO_MENSAL), linha_cabecalho)


def _montar_aba_pendencias(aba, dados: DadosRelatorio) -> None:
    aba.title = "Pendências"
    _aplicar_cabecalho(aba, dados, "Pendências")
    aba.append(_CABECALHO_PENDENCIAS)
    linha_cabecalho = aba.max_row

    for pendencia in sorted(
        dados.resultado.pendencias,
        key=lambda p: (p.nome_funcionario, p.data or datetime.min.date()),
    ):
        aba.append((
            pendencia.nome_funcionario,
            pendencia.data.strftime("%d/%m/%Y") if pendencia.data else "—",
            pendencia.tipo.value,
            pendencia.descricao,
            "Resolvida" if pendencia.resolvida else "Em aberto",
            pendencia.observacoes,
        ))

    linha_final = aba.max_row
    _aplicar_rodape(aba)
    _estilizar_titulo(aba, len(_CABECALHO_PENDENCIAS))
    _estilizar_tabela(aba, linha_cabecalho, linha_final, len(_CABECALHO_PENDENCIAS))
    _preparar_impressao(aba, len(_CABECALHO_PENDENCIAS), linha_cabecalho)


def _montar_aba_informacoes(aba, dados: DadosRelatorio) -> None:
    aba.title = "Informações do Processamento"
    _aplicar_cabecalho(aba, dados, "Informações do Processamento")
    linha_inicio_pares = aba.max_row + 1

    linhas = (
        ("Nome da empresa", dados.nome_empresa or "—"),
        ("Competência", dados.competencia_texto),
        ("Arquivo original utilizado", dados.arquivo_original),
        ("Quantidade de funcionários", dados.resumo_geral.funcionarios_processados),
        ("Quantidade de dias processados", dados.estatisticas.total_dias_processados),
        ("Quantidade de pendências", dados.resumo_geral.total_pendencias),
        ("Data do processamento", dados.data_processamento),
        ("Hora do processamento", dados.hora_processamento),
        ("Versão do sistema", VERSAO),
        ("Desenvolvedor", DESENVOLVEDOR),
    )
    for rotulo, valor in linhas:
        aba.append((rotulo, valor))

    aba.append(())
    linha_titulo_estatisticas = aba.max_row + 1
    aba.append(("Estatísticas da Competência (Cap. 11.10)",))
    aba.append(("Funcionários ativos", dados.estatisticas.funcionarios_ativos))
    aba.append(("Funcionários sem nenhuma batida", dados.estatisticas.funcionarios_sem_batidas))

    aba.append(())
    linha_titulo_pendencias = aba.max_row + 1
    aba.append(("Pendências por tipo",))
    for tipo, quantidade in dados.estatisticas.pendencias_por_tipo.items():
        aba.append((tipo.value, quantidade))

    aba.append(())
    linha_titulo_justificativas = aba.max_row + 1
    aba.append(("Distribuição das Justificativas",))
    for justificativa, quantidade in dados.estatisticas.distribuicao_justificativas.items():
        aba.append((justificativa, quantidade))

    linha_fim_pares = aba.max_row
    _aplicar_rodape(aba)

    from openpyxl.styles import Font

    _estilizar_titulo(aba, 2)
    _estilizar_pares_chave_valor(aba, linha_inicio_pares, linha_fim_pares)
    for linha in (linha_titulo_estatisticas, linha_titulo_pendencias, linha_titulo_justificativas):
        aba.cell(row=linha, column=1).font = Font(bold=True, italic=True, color="1F4E78")
    _preparar_impressao(aba, 2, None)


def _montar_aba_dashboard(aba, dados: DadosRelatorio) -> None:
    """
    Aba "Dashboard" (Etapa 11, v2.0) — indicadores, ranking e
    distribuição da competência, com gráficos nativos do Excel
    (`openpyxl.chart.BarChart`/`PieChart` — decisão confirmada com o
    usuário: sem nenhuma dependência nova, instalador continua leve).
    Nenhum valor é recalculado aqui: mesma fonte agregada de
    `dados.resumo_geral`/`dados.resultado.resumos_mensais`/
    `dados.resultado.pendencias` já usada pelas demais abas e pela
    Tela Dashboard (Etapa 10) in-app.
    """
    aba.title = "Dashboard"
    _aplicar_cabecalho(aba, dados, "Dashboard")

    resultado = dados.resultado
    resumo_geral = dados.resumo_geral
    resumos = resultado.resumos_mensais
    dias_processados = len({
        dia.data for funcionario in resultado.funcionarios_processados
        for dia in funcionario.dias
    })

    aba.append(("Indicadores Gerais",))
    linha_titulo_indicadores = aba.max_row
    linha_inicio_indicadores = aba.max_row + 1
    aba.append(("Total de Funcionários", resumo_geral.funcionarios_processados))
    aba.append(("Horas Trabalhadas", formatar_minutos(resumo_geral.horas_trabalhadas_min)))
    aba.append(("Horas Extras", formatar_minutos(resumo_geral.horas_extras_min)))
    aba.append(("Horas Negativas", formatar_minutos(resumo_geral.horas_negativas_min)))
    aba.append(("Banco de Horas (Saldo Geral)", formatar_minutos(resumo_geral.saldo_geral_min)))
    aba.append(("Pendências", resumo_geral.total_pendencias))
    aba.append(("Dias Processados", dias_processados))
    aba.append(("Competência", dados.competencia_texto))
    linha_fim_indicadores = aba.max_row
    aba.append(())

    aba.append(("Ranking de Horas Extras por Funcionário",))
    linha_titulo_ranking_extras = aba.max_row
    aba.append(("Funcionário", "Horas Extras (min)"))
    linha_cabecalho_ranking_extras = aba.max_row
    ranking_extras = sorted(
        (r for r in resumos if r.horas_extras_min > 0),
        key=lambda r: r.horas_extras_min, reverse=True,
    )[:10]
    for registro in ranking_extras:
        aba.append((registro.nome, registro.horas_extras_min))
    linha_fim_ranking_extras = aba.max_row
    aba.append(())

    aba.append(("Ranking de Horas Negativas por Funcionário",))
    linha_titulo_ranking_negativas = aba.max_row
    aba.append(("Funcionário", "Horas Negativas (min)"))
    linha_cabecalho_ranking_negativas = aba.max_row
    ranking_negativas = sorted(
        (r for r in resumos if r.horas_negativas_min > 0),
        key=lambda r: r.horas_negativas_min, reverse=True,
    )[:10]
    for registro in ranking_negativas:
        aba.append((registro.nome, registro.horas_negativas_min))
    linha_fim_ranking_negativas = aba.max_row
    aba.append(())

    aba.append(("Distribuição do Banco de Horas",))
    linha_titulo_distribuicao = aba.max_row
    aba.append(("Situação", "Quantidade de Funcionários"))
    linha_cabecalho_distribuicao = aba.max_row
    aba.append(("Positivo", sum(1 for r in resumos if r.saldo_final_min > 0)))
    aba.append(("Negativo", sum(1 for r in resumos if r.saldo_final_min < 0)))
    aba.append(("Zerado", sum(1 for r in resumos if r.saldo_final_min == 0)))
    linha_fim_distribuicao = aba.max_row

    _aplicar_rodape(aba)

    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font

    _estilizar_titulo(aba, 2)
    _estilizar_pares_chave_valor(aba, linha_inicio_indicadores, linha_fim_indicadores)
    for linha_titulo, linha_cabecalho, linha_fim in (
        (linha_titulo_ranking_extras, linha_cabecalho_ranking_extras, linha_fim_ranking_extras),
        (linha_titulo_ranking_negativas, linha_cabecalho_ranking_negativas,
         linha_fim_ranking_negativas),
        (linha_titulo_distribuicao, linha_cabecalho_distribuicao, linha_fim_distribuicao),
    ):
        aba.cell(row=linha_titulo, column=1).font = Font(bold=True, italic=True, color="1F4E78")
        _estilizar_tabela(aba, linha_cabecalho, linha_fim, 2)
    aba.cell(row=linha_titulo_indicadores, column=1).font = Font(
        bold=True, italic=True, color="1F4E78")

    if linha_fim_ranking_extras > linha_cabecalho_ranking_extras:
        grafico_extras = BarChart()
        grafico_extras.title = "Horas Extras por Funcionário (min)"
        grafico_extras.y_axis.title = "Minutos"
        grafico_extras.add_data(
            Reference(
                aba, min_col=2, min_row=linha_cabecalho_ranking_extras,
                max_row=linha_fim_ranking_extras),
            titles_from_data=True,
        )
        grafico_extras.set_categories(Reference(
            aba, min_col=1, min_row=linha_cabecalho_ranking_extras + 1,
            max_row=linha_fim_ranking_extras))
        aba.add_chart(grafico_extras, f"E{linha_titulo_indicadores}")

    if linha_fim_ranking_negativas > linha_cabecalho_ranking_negativas:
        grafico_negativas = BarChart()
        grafico_negativas.title = "Horas Negativas por Funcionário (min)"
        grafico_negativas.y_axis.title = "Minutos"
        grafico_negativas.add_data(
            Reference(
                aba, min_col=2, min_row=linha_cabecalho_ranking_negativas,
                max_row=linha_fim_ranking_negativas),
            titles_from_data=True,
        )
        grafico_negativas.set_categories(Reference(
            aba, min_col=1, min_row=linha_cabecalho_ranking_negativas + 1,
            max_row=linha_fim_ranking_negativas))
        aba.add_chart(grafico_negativas, f"E{linha_titulo_indicadores + 18}")

    grafico_pizza = PieChart()
    grafico_pizza.title = "Distribuição do Banco de Horas"
    grafico_pizza.add_data(
        Reference(
            aba, min_col=2, min_row=linha_cabecalho_distribuicao,
            max_row=linha_fim_distribuicao),
        titles_from_data=True,
    )
    grafico_pizza.set_categories(Reference(
        aba, min_col=1, min_row=linha_cabecalho_distribuicao + 1, max_row=linha_fim_distribuicao))
    aba.add_chart(grafico_pizza, f"E{linha_titulo_indicadores + 36}")

    _preparar_impressao(aba, 2, None)


def gerar_relatorio_excel_geral(dados: DadosRelatorio, caminho: Path) -> Path:
    """
    Gera o arquivo .xlsx com as 5 abas do Relatório Geral (Cap. 11.4 +
    Etapa 11 v2.0: nova aba "Dashboard" com indicadores, ranking,
    distribuição e gráficos nativos). Nunca toca a planilha original
    (Cap. 11.1/11.2) — só escreve um arquivo novo. Levanta
    `RelatorioBloqueadoError` se houver qualquer pendência em aberto
    (Cap. 9.1/11.11).
    """
    _garantir_sem_bloqueio(dados.resultado)
    import openpyxl

    livro = openpyxl.Workbook()
    _montar_aba_relatorio_diario(livro.active, dados)
    _montar_aba_resumo_mensal(livro.create_sheet(), dados)
    _montar_aba_pendencias(livro.create_sheet(), dados)
    _montar_aba_informacoes(livro.create_sheet(), dados)
    _montar_aba_dashboard(livro.create_sheet(), dados)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    livro.save(caminho)
    log.info("Relatório geral gerado: %s", caminho)
    return caminho


# ---------------------------------------------------------------------------
# Relatório Individual — 1 aba (Cap. 11.8)
# ---------------------------------------------------------------------------

def gerar_relatorio_excel_individual(
    dados: DadosRelatorio, funcionario: Funcionario, caminho: Path,
) -> Path:
    """
    Gera o .xlsx do Relatório Individual (Cap. 11.8) de um único
    funcionário. Levanta `RelatorioBloqueadoError` se houver qualquer
    pendência em aberto (Cap. 9.1/11.11) — em qualquer funcionário da
    competência, não só no funcionário deste relatório.
    """
    _garantir_sem_bloqueio(dados.resultado)
    import openpyxl

    resumo = montar_resumo_individual(funcionario, dados.resultado, dados.setores, dados.turnos)

    livro = openpyxl.Workbook()
    aba = livro.active
    aba.title = "Relatório Individual"
    _aplicar_cabecalho(aba, dados, f"Relatório Individual — {funcionario.nome_completo}")

    linha_inicio_info = aba.max_row + 1
    aba.append(("Nome", funcionario.nome_completo))
    aba.append(("Cargo", resumo.cargo))
    aba.append(("Setor", resumo.setor_nome))
    aba.append(("Turno", resumo.turno_nome))
    aba.append(("Competência", dados.competencia_texto))
    linha_fim_info = aba.max_row
    aba.append(())

    aba.append((
        "Data", "Dia da Semana", "Batidas", "Jornada Prevista", "Horas Trabalhadas",
        "Saldo", "Horas Extras", "Horas Negativas", "Situação", "Justificativa", "Observação",
    ))
    num_colunas_tabela = 11
    linha_cabecalho_tabela = aba.max_row
    for dia in sorted(funcionario.dias, key=lambda d: d.data):
        aba.append((
            dia.data.strftime("%d/%m/%Y"),
            dia.dia_semana,
            _texto_batidas(dia),
            formatar_minutos(dia.resultado.jornada_prevista_min),
            formatar_minutos(dia.resultado.horas_trabalhadas_min),
            formatar_minutos(dia.resultado.saldo_min),
            formatar_minutos(dia.resultado.horas_extras_min),
            formatar_minutos(dia.resultado.horas_negativas_min),
            dia.resultado.situacao.value,
            _justificativa_do_dia(dia),
            _observacao_do_dia(dia),
        ))
    linha_final_tabela = aba.max_row

    aba.append(())
    linha_titulo_resumo = aba.max_row + 1
    aba.append(("Resumo do Funcionário",))
    linha_inicio_resumo = aba.max_row + 1
    aba.append(("Horas Previstas", formatar_minutos(resumo.horas_previstas_min)))
    aba.append(("Horas Trabalhadas", formatar_minutos(resumo.horas_trabalhadas_min)))
    aba.append(("Horas Extras", formatar_minutos(resumo.horas_extras_min)))
    aba.append(("Horas Negativas", formatar_minutos(resumo.horas_negativas_min)))
    aba.append(("Saldo Final", formatar_minutos(resumo.saldo_final_min)))
    aba.append(("Pendências Resolvidas", resumo.quantidade_pendencias_resolvidas))
    aba.append(("Pendências Existentes", resumo.quantidade_pendencias_existentes))
    linha_fim_resumo = aba.max_row

    _aplicar_rodape(aba)

    from openpyxl.styles import Font

    _estilizar_titulo(aba, num_colunas_tabela)
    _estilizar_pares_chave_valor(aba, linha_inicio_info, linha_fim_info)
    _estilizar_tabela(aba, linha_cabecalho_tabela, linha_final_tabela, num_colunas_tabela)
    _estilizar_pares_chave_valor(aba, linha_inicio_resumo, linha_fim_resumo)
    aba.cell(row=linha_titulo_resumo, column=1).font = Font(bold=True, italic=True, color="1F4E78")
    _preparar_impressao(aba, num_colunas_tabela, linha_cabecalho_tabela)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    livro.save(caminho)
    log.info("Relatório individual gerado: %s (%s)", caminho, funcionario.nome_completo)
    return caminho


# ---------------------------------------------------------------------------
# Histórico (Cap. 11.7 / 14) — nunca sobrescreve
# ---------------------------------------------------------------------------

def caminho_historico(config: Config, competencia: tuple[int, int], nome_base: str) -> Path:
    """
    Monta o caminho de destino no Histórico (Cap. 11.7/14): Ano/Mês,
    com sufixo incremental (_001, _002...) se já existir um arquivo com
    o mesmo nome — nunca sobrescreve um relatório existente.
    """
    from constantes import nome_mes

    mes, ano = competencia
    pasta_base = Path(config.configuracoes.get("pasta_historico") or "Historico")
    pasta_destino = pasta_base / str(ano) / nome_mes(mes)
    pasta_destino.mkdir(parents=True, exist_ok=True)

    candidato = pasta_destino / f"{nome_base}.xlsx"
    if not candidato.exists():
        return candidato

    indice = 1
    while True:
        candidato = pasta_destino / f"{nome_base}_{indice:03d}.xlsx"
        if not candidato.exists():
            return candidato
        indice += 1


# ---------------------------------------------------------------------------
# Histórico (Cap. 12.6/14) — Sprint 6: leitura e exclusão do que já foi
# exportado. Nenhuma competência já fechada é recalculada aqui — só se
# lê o que o próprio relatório .xlsx já gravou (Cap. 11.4, Aba 4) ou se
# apaga arquivos já existentes, sempre dentro da pasta de Histórico.
# ---------------------------------------------------------------------------

_ROTULOS_INFORMACOES = {
    "Nome da empresa": "nome_empresa",
    "Quantidade de funcionários": "quantidade_funcionarios",
    "Quantidade de pendências": "quantidade_pendencias",
    "Data do processamento": "data_processamento",
    "Hora do processamento": "hora_processamento",
}


def _ler_informacoes_processamento(caminho: Path) -> dict[str, str]:
    """
    Lê de volta a aba "Informações do Processamento" (Cap. 11.4, Aba 4)
    de um relatório já exportado — nenhum valor é recalculado, só lido
    do que já está gravado em disco. Retorna um dict vazio se o
    arquivo não existir, estiver corrompido/em uso, ou não tiver essa
    aba (ex.: um Relatório Individual, que não a possui).
    """
    import openpyxl

    valores: dict[str, str] = {}
    livro = None
    try:
        livro = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        if "Informações do Processamento" not in livro.sheetnames:
            return valores
        aba = livro["Informações do Processamento"]
        for linha in aba.iter_rows(values_only=True):
            if not linha or linha[0] not in _ROTULOS_INFORMACOES:
                continue
            chave = _ROTULOS_INFORMACOES[linha[0]]
            valor = linha[1] if len(linha) > 1 else None
            valores[chave] = str(valor) if valor is not None else "—"
    except Exception as erro:  # pragma: no cover - proteção contra arquivo corrompido/em uso
        log.warning("Não foi possível ler informações do histórico em %s: %s", caminho, erro)
    finally:
        # Modo read_only mantém o arquivo aberto para leitura preguiçosa das
        # linhas — sem fechar explicitamente, o Windows mantém um lock nele
        # (ex.: bloquearia uma exclusão de histórico logo em seguida).
        if livro is not None:
            livro.close()
    return valores


def listar_competencias(config: Config) -> list[CompetenciaHistorico]:
    """
    Varre a pasta de Histórico (Cap. 12.6/14) e monta uma
    `CompetenciaHistorico` por Ano/Mês encontrado, com todos os
    arquivos `.xlsx` daquela competência (Geral e Individuais). As
    informações básicas vêm do primeiro relatório que tiver a aba
    "Informações do Processamento" — sem recalcular nada.
    """
    from constantes import MESES

    pasta_base = Path(config.configuracoes.get("pasta_historico") or "Historico")
    competencias: list[CompetenciaHistorico] = []
    if not pasta_base.exists():
        return competencias

    nomes_meses_invertido = {nome: numero for numero, nome in MESES.items()}

    try:
        pastas_ano = sorted(p for p in pasta_base.iterdir() if p.is_dir())
    except OSError as erro:
        log.warning("Não foi possível listar a pasta de Histórico %s: %s", pasta_base, erro)
        return competencias

    for pasta_ano in pastas_ano:
        if not pasta_ano.name.isdigit():
            continue
        ano = int(pasta_ano.name)

        try:
            pastas_mes = sorted(p for p in pasta_ano.iterdir() if p.is_dir())
        except OSError as erro:
            log.warning("Não foi possível listar %s: %s", pasta_ano, erro)
            continue

        for pasta_mes in pastas_mes:
            mes = nomes_meses_invertido.get(pasta_mes.name)
            if mes is None:
                continue

            try:
                caminhos_xlsx = sorted(
                    caminho for caminho in pasta_mes.glob("*.xlsx")
                    if not caminho.name.startswith("~$")  # lock do Excel, não um relatório
                )
                if not caminhos_xlsx:
                    continue

                competencia = CompetenciaHistorico(
                    ano=ano, mes=mes, competencia_texto=f"{pasta_mes.name}/{ano}",
                    pasta=pasta_mes,
                    arquivos=[
                        ArquivoHistorico(
                            caminho=caminho, nome=caminho.name,
                            tamanho_bytes=caminho.stat().st_size,
                            modificado_em=datetime.fromtimestamp(caminho.stat().st_mtime),
                        )
                        for caminho in caminhos_xlsx
                    ],
                )
            except OSError as erro:
                log.warning("Não foi possível ler os arquivos de %s: %s", pasta_mes, erro)
                continue

            for caminho in caminhos_xlsx:
                informacoes = _ler_informacoes_processamento(caminho)
                if informacoes:
                    for campo, valor in informacoes.items():
                        setattr(competencia, campo, valor)
                    break

            competencias.append(competencia)

    competencias.sort(key=lambda c: (c.ano, c.mes), reverse=True)
    return competencias


def excluir_historico(config: Config, competencia: CompetenciaHistorico) -> None:
    """
    Exclui todos os relatórios de uma competência (Cap. 12.6) — a
    pasta do Mês inteira. Nunca apaga nada fora da pasta de Histórico
    configurada (Cap. 11.7/14): recusa a operação se o alvo não for
    de fato uma subpasta dela, como proteção extra contra remover algo
    por engano.
    """
    import shutil

    pasta_base = Path(config.configuracoes.get("pasta_historico") or "Historico").resolve()
    pasta_alvo = competencia.pasta.resolve()
    if pasta_base not in pasta_alvo.parents:
        raise ValueError(
            f"Recusando excluir \"{pasta_alvo}\": não está dentro da pasta de "
            "Histórico configurada."
        )

    shutil.rmtree(pasta_alvo)
    log.info("Histórico excluído: %s (%s)", competencia.competencia_texto, pasta_alvo)


# ---------------------------------------------------------------------------
# Exportadores (Cap. 11.12) — arquitetura pronta para múltiplos formatos
# ---------------------------------------------------------------------------

class ExportadorRelatorio(Protocol):
    """Interface comum a qualquer formato de exportação do relatório (Cap. 11.12)."""

    def exportar_geral(self, dados: DadosRelatorio, caminho: Path) -> Path:
        """Gera o Relatório Geral (Cap. 11.4) no caminho informado e retorna o caminho final."""
        ...

    def exportar_individual(
        self, dados: DadosRelatorio, funcionario: Funcionario, caminho: Path,
    ) -> Path:
        """Gera o Relatório Individual (Cap. 11.8) e retorna o caminho final."""
        ...


class ExportadorExcel:
    """Implementação Excel (.xlsx) do `ExportadorRelatorio` (Cap. 11.12) — formato principal."""

    def exportar_geral(self, dados: DadosRelatorio, caminho: Path) -> Path:
        return gerar_relatorio_excel_geral(dados, caminho)

    def exportar_individual(
        self, dados: DadosRelatorio, funcionario: Funcionario, caminho: Path,
    ) -> Path:
        return gerar_relatorio_excel_individual(dados, funcionario, caminho)


class ExportadorPDF:
    """
    Implementação PDF do `ExportadorRelatorio` (Cap. 11.12) — arquitetura
    pronta, geração ainda não implementada nesta sprint: requer escolher
    e adicionar ao projeto uma biblioteca de PDF (ex.: reportlab ou
    fpdf2), o que não está no requirements.txt atual e não foi aprovado.
    Reservado para uma sprint futura.
    """

    def exportar_geral(self, dados: DadosRelatorio, caminho: Path) -> Path:
        raise NotImplementedError(
            "Exportação em PDF ainda não implementada — arquitetura reservada (Cap. 11.12)."
        )

    def exportar_individual(
        self, dados: DadosRelatorio, funcionario: Funcionario, caminho: Path,
    ) -> Path:
        raise NotImplementedError(
            "Exportação em PDF ainda não implementada — arquitetura reservada (Cap. 11.12)."
        )
