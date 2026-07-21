"""
exportacao.py
--------------
Exportação universal de tabelas (Sprint 1, v2.1) — Excel, CSV e PDF
genéricos para qualquer listagem do sistema (Funcionários, Setores,
Histórico, Competências, Pendências), usados por
`componentes.BotaoExportar`.

Distinto de `relatorio.py`: aquele módulo gera o Relatório Geral/
Individual da competência, com layout específico (Cap. 11). Este
módulo gera uma tabela genérica simples (título + cabeçalho + linhas)
a partir de qualquer lista de registros já formatados pela tela
chamadora — nenhuma regra de negócio, só formatação de exportação.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from config import Config
from constantes import APP_NOME

_COR_CABECALHO = "1F4E78"


def caminho_exportacao(config: Config, categoria: str, nome_base: str, extensao: str) -> Path:
    """
    Monta o caminho de destino de uma exportação genérica de tabela
    (Sprint 1, v2.1) — `Historico/Exportações/<categoria>/`, com
    sufixo incremental se já existir um arquivo com o mesmo nome no
    mesmo dia (mesmo espírito de `relatorio.caminho_historico`, mas
    sem exigir competência — Funcionários/Setores são cadastros
    globais, não por mês/ano).
    """
    pasta_base = Path(config.configuracoes.get("pasta_historico") or "Historico")
    pasta_destino = pasta_base / "Exportações" / categoria
    pasta_destino.mkdir(parents=True, exist_ok=True)

    data_texto = date.today().strftime("%Y-%m-%d")
    candidato = pasta_destino / f"{nome_base}_{data_texto}.{extensao}"
    if not candidato.exists():
        return candidato

    indice = 1
    while True:
        candidato = pasta_destino / f"{nome_base}_{data_texto}_{indice:03d}.{extensao}"
        if not candidato.exists():
            return candidato
        indice += 1


def exportar_excel_simples(
    caminho: Path, titulo: str, colunas: list[str], linhas: list[tuple],
) -> None:
    """Gera um .xlsx simples (título + cabeçalho + linhas) de uma tabela genérica."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    livro = openpyxl.Workbook()
    aba = livro.active
    aba.title = (titulo or "Exportação")[:31]

    aba.append((titulo,))
    aba.append((APP_NOME,))
    aba.append(())
    aba.append(tuple(colunas))
    linha_cabecalho = aba.max_row
    for linha in linhas:
        aba.append(linha)

    aba.cell(row=1, column=1).font = Font(bold=True, size=14, color=_COR_CABECALHO)
    aba.cell(row=2, column=1).font = Font(italic=True, color="595959")
    for coluna in range(1, len(colunas) + 1):
        celula = aba.cell(row=linha_cabecalho, column=coluna)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=_COR_CABECALHO)

    for coluna in range(1, len(colunas) + 1):
        maior = max(
            (len(str(aba.cell(row=linha, column=coluna).value or ""))
             for linha in range(linha_cabecalho, aba.max_row + 1)),
            default=10,
        )
        aba.column_dimensions[get_column_letter(coluna)].width = min(maior + 4, 50)

    aba.freeze_panes = aba.cell(row=linha_cabecalho + 1, column=1)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    livro.save(caminho)


def exportar_csv(caminho: Path, colunas: list[str], linhas: list[tuple]) -> None:
    """Gera um .csv com separador ';' (padrão Excel PT-BR) e BOM UTF-8."""
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with open(caminho, "w", newline="", encoding="utf-8-sig") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow(colunas)
        escritor.writerows(linhas)


def exportar_pdf_simples(
    caminho: Path, titulo: str, colunas: list[str], linhas: list[tuple],
) -> None:
    """Gera um .pdf simples (título + tabela) via ReportLab, paisagem A4."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    caminho.parent.mkdir(parents=True, exist_ok=True)
    documento = SimpleDocTemplate(
        str(caminho), pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(titulo or "Exportação", estilos["Title"]),
        Paragraph(APP_NOME, estilos["Normal"]),
        Spacer(1, 12),
    ]

    dados = [colunas] + [[str(valor) for valor in linha] for linha in linhas]
    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_COR_CABECALHO}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabela)
    documento.build(elementos)
